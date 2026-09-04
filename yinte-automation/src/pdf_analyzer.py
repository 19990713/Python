"""
PDF 图纸分析 — 统计页面尺寸，生成 Excel
基于 图纸统计.py 的核心逻辑
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional

from pypdf import PdfReader

PT_TO_MM = 0.352778

STANDARD_SIZES = {
    "A0": (841, 1189), "A1": (594, 841), "A2": (420, 594),
    "A3": (297, 420), "A4": (210, 297), "A5": (148, 210),
    "B0": (1000, 1414), "B1": (707, 1000), "B2": (500, 707),
    "B3": (353, 500), "B4": (250, 353), "B5": (176, 250),
}

TOLERANCE_MM = 5
EXTEND_THRESHOLD_MM = 20  # 长边超出标准 20mm 以上视为+

# 尺寸排序权重：A4 在最上面，其余由小到大（按短边尺寸）
# +版本排在对应标准尺寸之后（权重 +0.5）
_SIZE_ORDER = {
    "A4": 0,    # A4 始终在最上面
    "A5": 1, "B5": 2,
    "B4": 3,
    "A3": 4, "B3": 5,
    "A2": 6, "B2": 7,
    "A1": 8, "B1": 9,
    "A0": 10, "B0": 11,
}


def get_size_sort_key(size_name: str) -> float:
    """返回排序权重：A4 最前 → 小尺寸 → 大尺寸 → 其他。
    +版本（+后缀）排在对应标准尺寸之后（权重 +0.5）。"""
    base_name = size_name.rstrip("+")
    is_extended = size_name.endswith("+")
    base_order = _SIZE_ORDER.get(base_name, 99)
    return base_order + (0.5 if is_extended else 0)


def get_standard_size_name(width_mm: float, height_mm: float) -> Tuple[str, str]:
    """
    根据实际尺寸（mm）判断标准图纸尺寸名称及匹配精度。
    返回 (尺寸名称, 匹配类型)。

    匹配类型: "精确匹配" | "+" | "近似匹配" | "非标准" | "非标准(长边偏短)"
    使用三级匹配策略:
      1) 短边容差匹配 (±TOLERANCE_MM)
      2) 比例容差匹配 (短边偏差 ≤ 5%)
      3) 最近短边匹配
    """
    short_side = min(width_mm, height_mm)
    long_side = max(width_mm, height_mm)

    # ----- 第一轮：短边容差匹配 -----
    candidates = []
    for name, (sw, sl) in STANDARD_SIZES.items():
        short_diff = abs(short_side - sw)
        if short_diff <= TOLERANCE_MM:
            long_diff = long_side - sl
            candidates.append((name, sw, sl, short_diff, long_diff))

    if not candidates:
        # ----- 第二轮：比例容差匹配 (短边偏差 ≤ 5%) -----
        for name, (sw, sl) in STANDARD_SIZES.items():
            short_diff_pct = abs(short_side - sw) / sw
            if short_diff_pct <= 0.05:
                long_diff = long_side - sl
                if long_diff >= -TOLERANCE_MM:
                    if long_diff > EXTEND_THRESHOLD_MM:
                        return f"{name}+", "+"
                    else:
                        return name, "精确匹配"

        # ----- 第三轮：最近短边匹配 -----
        min_short_diff = float('inf')
        best = None
        for name, (sw, sl) in STANDARD_SIZES.items():
            short_diff = abs(short_side - sw)
            if short_diff < min_short_diff:
                min_short_diff = short_diff
                best = (name, sw, sl, short_diff, long_side - sl)
        if best is None:
            return "其他", "非标准"

        best_name, best_sw, best_sl, best_short_diff, best_long_diff = best
        if best_long_diff > EXTEND_THRESHOLD_MM:
            return f"{best_name}+", "+"
        return best_name, "近似匹配"

    # 在短边匹配的候选中，按长边差值绝对值排序
    candidates.sort(key=lambda x: abs(x[4]))
    best_name, best_sw, best_sl, best_short_diff, best_long_diff = candidates[0]

    if best_long_diff <= EXTEND_THRESHOLD_MM and best_long_diff >= -TOLERANCE_MM:
        return best_name, "精确匹配"
    elif best_long_diff > EXTEND_THRESHOLD_MM:
        return f"{best_name}+", "+"
    else:
        return best_name, "非标准(长边偏短)"


# 向后兼容别名
def get_size_name(w_mm: float, h_mm: float) -> str:
    """根据宽高判断标准图纸尺寸名称（向后兼容）。
    新代码建议用 get_standard_size_name()。"""
    name, _ = get_standard_size_name(w_mm, h_mm)
    return name


def analyze_pdfs(paths: List[Path], progress_callback=None) -> Tuple[Dict[str, int], int, List[Dict]]:
    """
    分析 PDF 文件列表，返回 (尺寸统计, 总页数, 详细数据)。

    详细数据列表中每项包含:
      文件名, 相对路径, 页码, 总页数, 宽度(mm), 高度(mm),
      页面方向, 标准尺寸分类, 匹配精度
    """
    size_stats: Dict[str, int] = {}
    total_pages = 0
    total = len(paths)
    detail_data: List[Dict] = []

    for i, pdf_path in enumerate(paths):
        try:
            reader = PdfReader(pdf_path)
            pages = len(reader.pages)
            total_pages += pages

            for page_num, page in enumerate(reader.pages, start=1):
                w_mm = float(page.mediabox.width) * PT_TO_MM
                h_mm = float(page.mediabox.height) * PT_TO_MM
                size_name, match_type = get_standard_size_name(w_mm, h_mm)
                size_stats[size_name] = size_stats.get(size_name, 0) + 1

                detail_data.append({
                    "文件名": pdf_path.name,
                    "相对路径": str(pdf_path),
                    "页码": page_num,
                    "总页数": pages,
                    "宽度(mm)": round(w_mm, 2),
                    "高度(mm)": round(h_mm, 2),
                    "页面方向": "横向" if w_mm > h_mm else ("纵向" if h_mm > w_mm else "方形"),
                    "标准尺寸分类": size_name,
                    "匹配精度": match_type,
                })

            if progress_callback:
                progress_callback(i + 1, total, pdf_path.name)

        except Exception as e:
            print(f"  跳过 {pdf_path.name}: {e}")

    return size_stats, total_pages, detail_data


def generate_excel(
    size_stats: Dict[str, int],
    total_pages: int,
    pdf_count: int,
    subtype: str,
    copies: int,
    output_path: str,
    detail_data: Optional[List[Dict]] = None,
) -> str:
    """
    生成页面尺寸统计 Excel。
    包含两个工作表: "统计汇总"（在前）和 "详细数据"（在后）。
    格式与 main.py 期望的输入一致。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ===== 统计汇总工作表 =====
    ws = wb.create_sheet("统计汇总", 0)

    # 标题行
    ws.append(["图纸尺寸统计汇总"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([f"统计日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"PDF文件总数: {pdf_count}"])
    ws.append([f"总页数: {total_pages}"])

    # B4 写子项类型
    subtype_cell = ws.cell(row=4, column=2)
    subtype_cell.value = subtype
    subtype_cell.font = Font(bold=True, color="FF0000")

    ws.append([])  # 第5行空

    # 表头（第6行）
    headers = ["尺寸分类", "页数", "份数", "数量", "占比(%)"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col in range(1, 6):
        cell = ws.cell(row=6, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 颜色池（每个尺寸分类一个颜色）
    color_pool = [
        "FFB3B3", "B3D9FF", "B3FFB3", "FFFFB3", "FFB3FF",
        "B3FFFF", "FFD9B3", "D9B3FF", "B3B3FF", "FFB3D9",
        "C4E0F9", "F9C4C4", "C4F9C4", "F9F9C4", "C4C4F9"
    ]
    size_color_map = {}
    idx = 0
    for size_name in sorted(size_stats.keys()):
        size_color_map[size_name] = color_pool[idx % len(color_pool)]
        idx += 1

    # 数据（按尺寸排序）
    sorted_stats = sorted(size_stats.items(), key=lambda x: get_size_sort_key(x[0]))
    for i, (size_name, count) in enumerate(sorted_stats):
        row = 7 + i
        ratio = (count / total_pages * 100) if total_pages > 0 else 0
        ws.append([size_name, count, copies, f"=B{row}*C{row}", round(ratio, 2)])
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=size_color_map[size_name],
                                    end_color=size_color_map[size_name], fill_type="solid")
            cell.alignment = center
            cell.border = border

    # 总计行
    total_row = 7 + len(sorted_stats)
    ws.append(["总计", f"=SUM(B7:B{total_row-1})", "", f"=SUM(D7:D{total_row-1})", 100.0])
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = center
        cell.border = border

    # 列宽
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    # ===== 详细数据工作表 =====
    if detail_data:
        ws_detail = wb.create_sheet("详细数据", 1)
        fieldnames = ["文件名", "相对路径", "页码", "总页数",
                      "宽度(mm)", "高度(mm)", "页面方向",
                      "标准尺寸分类", "匹配精度"]
        ws_detail.append(fieldnames)

        for col_idx, header in enumerate(fieldnames, start=1):
            cell = ws_detail.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = center

        for row_data in detail_data:
            ws_detail.append([
                row_data["文件名"],
                row_data["相对路径"],
                row_data["页码"],
                row_data["总页数"],
                row_data["宽度(mm)"],
                row_data["高度(mm)"],
                row_data["页面方向"],
                row_data["标准尺寸分类"],
                row_data["匹配精度"],
            ])

        # 自适应列宽
        for col_idx, col_name in enumerate(fieldnames, start=1):
            max_len = len(col_name)
            for row_data in detail_data:
                val = str(row_data.get(col_name, ""))
                if len(val) > max_len:
                    max_len = len(val)
            adjusted_len = min(max_len + 2, 35)
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = adjusted_len

    wb.save(output_path)
    return output_path


def collect_pdfs(args: List[str]) -> Tuple[List[Path], Optional[Path]]:
    """从命令行参数收集 PDF 文件。"""
    pdfs = []
    base_folder = None

    for arg in args:
        p = Path(arg)
        if p.is_file() and p.suffix.lower() == ".pdf":
            pdfs.append(p)
        elif p.is_dir():
            pdfs.extend(p.rglob("*.pdf"))
            if base_folder is None:
                base_folder = p

    # 去重
    seen = set()
    unique = []
    for p in pdfs:
        if p not in seen:
            seen.add(p)
            unique.append(p)

    if not base_folder and unique:
        base_folder = unique[0].parent

    return unique, base_folder
