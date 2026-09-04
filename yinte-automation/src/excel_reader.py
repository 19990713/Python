"""
Excel 读取模块 — 从 .xlsx/.xls 文件读取数据，返回字典列表。
"""

import logging
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd

logger = logging.getLogger("yinte.excel")

# 自动检测时需要匹配的关键列名
_DETECT_KEYWORDS = ["尺寸分类", "数量", "份数", "页数", "子项", "备注"]


def auto_detect_header_row(
    filepath: str,
    sheet_name: str | int = 0,
    target_columns: List[str] | None = None,
    max_scan_rows: int = 20,
) -> int:
    """
    自动扫描 Excel 前 N 行，找到包含目标列名最多的那一行作为标题行。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名或索引
        target_columns: 要匹配的目标列名列表，默认使用 _DETECT_KEYWORDS
        max_scan_rows: 最多扫描的行数

    Returns:
        1-based 标题行号，找不到返回 0
    """
    if target_columns is None:
        target_columns = _DETECT_KEYWORDS

    try:
        df_raw = pd.read_excel(
            filepath, sheet_name=sheet_name,
            header=None, dtype=str, nrows=max_scan_rows
        )
    except Exception as e:
        logger.warning(f"自动检测标题行失败（无法读取Excel）: {e}")
        return 0

    if df_raw.empty:
        return 0

    best_row = 0
    best_score = 0
    best_matches = []

    for row_idx, row in df_raw.iterrows():
        # 将该行每个非空单元格转为字符串
        cell_values = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip()]
        if not cell_values:
            continue

        # 子串匹配：单元格值包含目标列名即可（兼容"尺寸分类(mm)"等变体）
        matches = [tc for tc in target_columns if any(tc in cv for cv in cell_values)]
        score = len(matches)

        if score > best_score:
            best_score = score
            best_row = row_idx
            best_matches = matches

    # 至少命中 2 个目标列才认为有效（尺寸分类 + 数量/份数/页数任一）
    if best_score >= 2:
        header_1based = best_row + 1  # pandas row_idx 是 0-based
        logger.info(
            f"自动检测到标题行: 第 {header_1based} 行, "
            f"匹配列: {best_matches} ({best_score}/{len(target_columns)})"
        )
        return header_1based

    logger.info(f"自动检测标题行失败（最佳匹配仅 {best_score} 列: {best_matches}）")
    return 0


def read_excel(
    filepath: str,
    sheet_name: str | int = 0,
    header_row: int = 0,
    target_columns: List[str] | None = None,
) -> List[Dict[str, Any]]:
    """
    读取 Excel 文件，返回每行数据的字典列表。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名或索引（默认第一个表）
        header_row: 标题行号，从1开始（1=第一行）。
                    0 或负数 = 自动检测（扫描前20行找目标列名）。
        target_columns: 自动检测时匹配的列名，默认使用内置关键词。

    Returns:
        [{"列名": 值, ...}, ...]
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    # 自动检测标题行
    if header_row <= 0:
        detected = auto_detect_header_row(
            filepath, sheet_name=sheet_name,
            target_columns=target_columns,
        )
        if detected > 0:
            header_row = detected
        else:
            # 回退：默认用第6行
            header_row = 6
            logger.warning(f"自动检测失败，回退使用第 {header_row} 行作为标题行")

    logger.info(f"读取 Excel: {filepath} (工作表: {sheet_name}, 标题行: {header_row})")

    header_0based = header_row - 1  # 转为pandas的0-based

    df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str, header=header_0based)

    # 去掉全空行
    df = df.dropna(how="all")

    # 去掉全空列
    df = df.dropna(axis=1, how="all")

    if df.empty:
        raise ValueError("Excel 中没有数据")

    # 清理列名（去首尾空格，处理Unnamed列）
    df.columns = [str(c).strip() for c in df.columns]

    # 将 NaN 替换为空字符串
    df = df.fillna("")

    records = df.to_dict(orient="records")
    # 去除首尾空格
    records = [{k: str(v).strip() for k, v in row.items()} for row in records]

    logger.info(f"读取到 {len(records)} 条数据, {len(df.columns)} 列: {list(df.columns)}")

    return records


def read_cell(filepath: str, sheet_name: str | int = 0, cell: str = "B4") -> str:
    """
    用 openpyxl 只读模式读取 Excel 中指定单元格的值（避免 pandas 整表重读）。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名或索引（默认第一个表）
        cell: 单元格地址，如 "B4"

    Returns:
        单元格值（字符串）
    """
    from openpyxl import load_workbook

    path = Path(filepath)
    if not path.exists():
        return ""

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        return str(ws[cell].value or "").strip()
    finally:
        wb.close()


def validate_columns(records: List[Dict], field_configs: List[Dict]) -> List[str]:
    """
    校验配置中的 excel_column 是否在 Excel 中实际存在。

    Returns:
        缺失的列名列表（空列表 = 全部通过）
    """
    if not records:
        return []

    excel_cols = set(records[0].keys())
    required = [f["excel_column"] for f in field_configs]
    missing = [c for c in required if c not in excel_cols]

    if missing:
        logger.error(f"Excel 中缺少以下列: {missing}")
        logger.info(f"Excel 实际列名: {sorted(excel_cols)}")

    return missing
