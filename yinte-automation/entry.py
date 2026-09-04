#!/usr/bin/env python
"""
印特自动化 — 统一入口
  拖入 Excel → 直接运行印特录入
  拖入 PDF/文件夹 → 分析图纸 → 询问类型+份数 → 直接录入
"""

import sys
import yaml
from pathlib import Path

# EXE模式: 从EXE所在目录读取配置; 脚本模式: 从脚本目录读取
if getattr(sys, 'frozen', False):
    APP_DIR = Path(sys.executable).parent.resolve()
else:
    APP_DIR = Path(__file__).parent.resolve()
sys.path.insert(0, str(APP_DIR))

from src.logger import setup_logging
from src.pdf_analyzer import get_size_sort_key

# 有效图纸类型
VALID_SUBTYPES = {"蓝图", "彩图", "白图", "复图", "硫酸纸"}
SUBTYPE_MAP = {"1": "蓝图", "2": "彩图", "3": "白图", "4": "复图", "5": "硫酸纸"}


def _load_config():
    """加载 config.yaml：优先读 EXE 同目录的外部文件，找不到则用内嵌默认。"""
    # 1) 外部 config.yaml（与 EXE 同目录），用户可编辑覆盖
    external = APP_DIR / "config.yaml"
    if external.exists():
        with open(external, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        if config is not None:
            return config

    # 2) 内嵌 config.yaml（PyInstaller --add-data 打包的）
    if getattr(sys, 'frozen', False):
        import os
        bundled = Path(sys._MEIPASS) / "config.yaml"
        if bundled.exists():
            with open(bundled, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f)
            if config is not None:
                return config

    raise FileNotFoundError("config.yaml 未找到或为空")


def _read_cell_light(filepath: str, cell: str) -> str:
    """用 openpyxl 只读模式直接读取单个单元格，避免 pandas 整表重读。"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        return str(ws[cell].value or "").strip()
    finally:
        wb.close()


def _ask_subtype(prompt_hint: str = "") -> str:
    """询问用户图纸类型，直到输入有效值。"""
    prompt = f"图纸类型 蓝图(1)/彩图(2)/白图(3)/复图(4)/硫酸纸(5): "
    subtype = input(prompt).strip()
    subtype = SUBTYPE_MAP.get(subtype, subtype)
    while subtype not in VALID_SUBTYPES:
        subtype = input(f"请输入有效{prompt}").strip()
        subtype = SUBTYPE_MAP.get(subtype, subtype)
    return subtype


def main():
    args = sys.argv[1:]
    if not args:
        print("请将 Excel/PDF/文件夹 拖到这个图标上！")
        input("按回车退出...")
        sys.exit(1)

    logger = setup_logging(str(APP_DIR / "logs"))

    records = []
    config = None

    # 分类输入
    is_pdf_mode = False

    for arg in args:
        p = Path(arg)
        if not p.exists():
            print(f"路径不存在: {arg}")
            continue
        if p.suffix.lower() == ".pdf":
            is_pdf_mode = True
        elif p.is_dir():
            is_pdf_mode = True
        elif p.suffix.lower() in (".xlsx", ".xls"):
            is_pdf_mode = False
            break

    # ===== PDF/文件夹模式 =====
    if is_pdf_mode:
        from src.pdf_analyzer import analyze_pdfs, collect_pdfs

        all_pdfs, _ = collect_pdfs(args)
        if not all_pdfs:
            print("未找到 PDF 文件")
            input("按回车退出...")
            sys.exit(1)

        print(f"共 {len(all_pdfs)} 个 PDF 文件")
        print("正在分析图纸尺寸...\n")
        size_stats, total_pages, _detail = analyze_pdfs(all_pdfs)

        print(f"总页数: {total_pages}")
        for name, count in sorted(size_stats.items(), key=lambda x: x[1], reverse=True):
            print(f"  {name}: {count} 页")

        # 询问用户
        print()
        subtype = _ask_subtype()

        copies_str = input("份数 (默认1): ").strip()
        try:
            copies = int(copies_str) if copies_str else 1
        except ValueError:
            copies = 1

        # 转换为 records (与 Excel 格式一致)
        for size_name, pages in size_stats.items():
            records.append({
                "尺寸分类": size_name,
                "子项": subtype,
                "数量": str(pages * copies),
                "份数": str(copies),
                "页数": str(pages),
            })

        # 加载配置
        config = _load_config()
        config["default_subtype"] = subtype

    # ===== Excel 模式 =====
    else:
        config = _load_config()

        excel_file = None
        for arg in args:
            p = Path(arg)
            if p.suffix.lower() in (".xlsx", ".xls"):
                excel_file = p
                break

        if not excel_file:
            print("未找到 Excel 文件")
            input("按回车退出...")
            sys.exit(1)

        from src.excel_reader import read_excel

        excel_cfg = config.get("excel", {})
        header_row = excel_cfg.get("header_row", 0)  # 0 = 自动检测

        logger.info(f"读取: {excel_file}")
        records = read_excel(str(excel_file), sheet_name=0, header_row=header_row)

        # 兼容旧格式: "A1加长" → "A1+"
        for r in records:
            val = r.get("尺寸分类", "")
            if isinstance(val, str) and "加长" in val:
                r["尺寸分类"] = val.replace("加长", "+")

        # 校验必要列（数量列可选，引擎仅使用份数+页数）
        col_cfg = config.get("columns", {})
        size_col = col_cfg["尺寸分类"]
        required = [size_col, col_cfg["份数"], col_cfg["页数"]]
        # 数量列为可选（部分Excel没有单独的"数量"列）
        qty_col = col_cfg.get("数量", "数量")
        excel_cols = set(records[0].keys()) if records else set()
        missing = [r for r in required if r not in excel_cols]
        if missing:
            logger.error(f"Excel 缺少列: {missing}")
            logger.info(f"实际列: {sorted(excel_cols)}")
            logger.info("请在 config.yaml 的 excel.header_row 设置正确的标题行号")
            print(f"Excel 缺少必要列: {missing}")
            print(f"   实际列: {sorted(excel_cols)}")
            input("按回车退出...")
            sys.exit(1)

        # 过滤汇总行
        records = [r for r in records if r.get(size_col, "").strip() not in ("总计", "合计", "小计", "")]
        if not records:
            print("无数据")
            input("按回车退出...")
            sys.exit(1)

        # 用 openpyxl 轻量读取 B4 单元格（避免整表重读）
        subtype_cell = config.get("subtype_cell", "B4")
        cell_value = _read_cell_light(str(excel_file), subtype_cell)

        if cell_value in VALID_SUBTYPES:
            config["default_subtype"] = cell_value
            print(f"检测到图纸类型: {cell_value}")
        else:
            if cell_value:
                print(f"B4 单元格内容 '{cell_value}' 不是有效图纸类型")
            print()
            config["default_subtype"] = _ask_subtype()

        logger.info(f"子项: {config['default_subtype']}")

    # ===== 按尺寸排序 (A4→A3→A2→A2+→A1→A1+→A0→A0+) =====
    size_col = config["columns"]["尺寸分类"]
    pages_col = config["columns"]["页数"]
    copies_col = config["columns"]["份数"]
    records.sort(key=lambda r: get_size_sort_key(str(r.get(size_col, ""))))

    # ===== 运行印特录入 =====
    if not records:
        print("无数据")
        input("按回车退出...")
        sys.exit(1)

    print(f"\n共 {len(records)} 条待录入 (已按尺寸排序)")
    for r in records:
        print(f"  {r.get(size_col, '?')} | 页数={r.get(pages_col, '?')} 份数={r.get(copies_col, '?')}")

    # ===== 模式选择 =====
    print()
    print("请选择录入模式:")
    print("  [1] 新建工单 (自动打开新建工单窗口)")
    print("  [2] 补录工单 (手动打开已有工单，追加数据)")
    mode = input("请输入 1 或 2 (默认1): ").strip() or "1"
    is_append = (mode == "2")

    from src.automation import AutomationEngine
    engine = AutomationEngine(config)

    if is_append:
        # ===== 补录模式 =====
        print("\n请手动打开需要补录的工单窗口...")
        input("打开后按回车继续...")

        hwnd = engine.find_work_order_window()
        if hwnd is None:
            print("未找到工单窗口，请确认工单已打开")
            input("按回车退出...")
            sys.exit(1)

        # 连接到工单窗口（用标题模糊匹配 工作单）
        try:
            engine.window_title = "工作单"
            engine.connect()
        except Exception:
            print("无法连接到工单窗口，请重试")
            input("按回车退出...")
            sys.exit(1)

        if not engine.click_new_row(hwnd):
            print("无法定位到空白行")
            input("按回车退出...")
            sys.exit(1)

        result = engine.run(records, append_mode=True)
    else:
        # ===== 新建模式（原有流程）=====
        print("\n请确认印特主窗口已打开！")
        input("按回车自动打开新建工单并开始录入...")

        # 自动打开新建工单
        if not engine.open_new_work_order():
            print("无法自动打开新建工单，请手动打开后重试")
            input("按回车退出...")
            sys.exit(1)

        engine.connect()
        result = engine.run(records)

    logger.info(f"成功 {result['success']}, 失败 {result['failed']}")

    if result["failed"] > 0:
        print(f"\n{result['failed']} 条录入失败，请检查日志")
        input("按回车退出...")
        sys.exit(2)


if __name__ == "__main__":
    main()
